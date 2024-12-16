import json
from flask import Flask, request, render_template, make_response, abort, flash, redirect, Blueprint
import openpyxl
from openpyxl import Workbook
from io import BytesIO, StringIO
import csv

app = Flask(__name__)
app.secret_key = 'your_secret_key'
monitoring_mrch_bp = Blueprint('monitoring_mrch', __name__)

def load_column_mapping():
    try:
        with open('data/monitoring_MRCH.json', 'r', encoding='utf-8') as file:
            return json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        print("Error: 'data/monitoring_MRCH.json' file not found or invalid JSON.")
        return {}

COLUMN_MAPPING = load_column_mapping()

THRESHOLD_NUMERIC_DIFF = 10
THRESHOLD_PERCENTAGE_DIFF = 10

last_results = None
last_summary = None

def load_products_from_file(file):
    """Load products from an Excel file and map columns based on COLUMN_MAPPING."""
    products = {}
    row_count = 0

    try:
        workbook = openpyxl.load_workbook(file, data_only=True)

        for sheet in workbook.worksheets:
            for row in range(2, sheet.max_row + 1):
                product_id_cell = COLUMN_MAPPING.get('id')
                if not product_id_cell:
                    continue

                col_a_value = sheet.cell(row=row, column=product_id_cell).value

                if col_a_value is None or str(col_a_value).strip() == '':
                    continue

                product_details = {}
                for key, col in COLUMN_MAPPING.items():
                    cell_value = sheet.cell(row=row, column=col).value
                    value = cell_value if cell_value is not None else ''
                    product_details[key] = value

                product_id = product_details.get('id', '')
                if not product_id:
                    continue

                try:
                    price_per_kg_col = COLUMN_MAPPING.get('price_per_kg', 0)
                    price_per_kg = float(sheet.cell(row=row, column=price_per_kg_col).value or 0)
                    product_details['price'] = round(price_per_kg, 2)
                except (ValueError, TypeError):
                    product_details['price'] = 0.0

                products[product_id] = product_details
                row_count += 1

    except openpyxl.utils.exceptions.InvalidFileException:
        abort(400, "Invalid file format. Please upload a valid Excel file.")
    except Exception as e:
        print(f"Unexpected error: {e}")
        abort(500, "An unexpected error occurred while processing the file.")

    return products, row_count


def calculate_numeric_difference(price1, price2):
    """Calculate the price difference and percentage difference."""
    numeric_diff = round(price2 - price1, 2)
    percentage_diff = round((numeric_diff / price1) * 100, 2) if price1 != 0 else None
    return numeric_diff, percentage_diff


def compare_product_data(file1, file2):
    global last_results, last_summary

    products1, file1_row_count = load_products_from_file(file1)
    products2, file2_row_count = load_products_from_file(file2)

    all_product_ids = set(products1.keys()).union(set(products2.keys()))

    diff_results = []

    for product_id in all_product_ids:
        product_data1 = products1.get(product_id, {})
        product_data2 = products2.get(product_id, {})

        result = {'id': product_id}

        result['photo_last_week'] = product_data1.get('photo', '')
        result['promo_last_week'] = product_data1.get('promo', '')

        for key in COLUMN_MAPPING.keys():
            if key != 'price':
                result[key] = product_data2.get(key, '')

        price1 = product_data1.get('price')
        price2 = product_data2.get('price')

        try:
            price1 = float(price1) if price1 not in [None, ''] else None
        except (ValueError, TypeError):
            price1 = None

        try:
            price2 = float(price2) if price2 not in [None, ''] else None
        except (ValueError, TypeError):
            price2 = None

        result['price1'] = price1 if price1 is not None else ''
        result['price2'] = price2 if price2 is not None else ''

        if price1 is not None and price2 is not None:
            numeric_diff, percentage_diff = calculate_numeric_difference(price1, price2)
            result['numeric_diff'] = numeric_diff
            result['percentage_diff'] = percentage_diff

            result['significant_diff'] = (
                abs(numeric_diff) >= THRESHOLD_NUMERIC_DIFF or
                (percentage_diff is not None and abs(percentage_diff) >= THRESHOLD_PERCENTAGE_DIFF)
            )
        else:
            result['numeric_diff'] = None
            result['percentage_diff'] = None
            result['significant_diff'] = False

        diff_results.append(result)

    last_results = diff_results
    last_summary = {
        "file1_row_count": file1_row_count,
        "file2_row_count": file2_row_count,
        "total_changes": sum(1 for r in diff_results if r['significant_diff'])
    }

    return diff_results, last_summary


def get_significant_diff_results(results):
    """Filter results to include only those with significant differences."""
    if not results:
        return []
    significant_results = [
        result for result in results
        if result.get('significant_diff', False)
    ]
    return significant_results

@monitoring_mrch_bp.route('/result', methods=['GET', 'POST'])
def monitoring_MRCH():
    if request.method == 'POST':
        file1 = request.files.get('file1')
        file2 = request.files.get('file2')

        if not file1 or not file2 or not file1.filename.endswith('.xlsx') or not file2.filename.endswith('.xlsx'):
            return render_template('error.html', error="Please upload two valid Excel files (.xlsx).")

        try:
            results, summary = compare_product_data(file1, file2)
            significant_results = get_significant_diff_results(results)
            if not significant_results:
                return render_template('error.html', error="No significant differences found")
            return render_template('monitoring_MRCH/iteration.html', results=significant_results, summary=summary)
        except Exception as e:
            print(f"Error: {e}")
            return render_template('error.html', error="File processing error. Please try again.")

    return render_template('monitoring_MRCH/index.html')

@monitoring_mrch_bp.route('/edit_prices_MRCH', methods=['POST'])
def edit_prices_MRCH():
    global last_results, last_summary
    if last_results is None:
        abort(400, "No data available to edit")

    try:
        for result in last_results:
            product_id = result['id']
            promo_key = f"promo_{product_id}"
            if promo_key in request.form:
                new_promo = request.form[promo_key].strip()
                result['promo'] = new_promo

            price2_key = f"price2_{product_id}"
            if price2_key in request.form:
                new_price2 = request.form[price2_key].strip()
                if new_price2:
                    try:
                        new_price2_float = float(new_price2)
                        result['price2'] = new_price2_float
                        if 'price1' in result and result['price1'] != '':
                            price1 = float(result['price1'])
                            numeric_diff, percentage_diff = calculate_numeric_difference(price1, new_price2_float)
                            result['numeric_diff'] = numeric_diff
                            result['percentage_diff'] = percentage_diff
                        else:
                            result['numeric_diff'] = None
                            result['percentage_diff'] = None
                    except ValueError:
                        flash(f"Invalid price for product {product_id}", "danger")
                else:
                    result['price2'] = ''
                    result['numeric_diff'] = None
                    result['percentage_diff'] = None

        # Перераховуємо summary
        diff_items = [
            r for r in last_results
            if r['numeric_diff'] is not None and (
                abs(r['numeric_diff']) >= THRESHOLD_NUMERIC_DIFF or
                (r['percentage_diff'] is not None and abs(r['percentage_diff']) >= THRESHOLD_PERCENTAGE_DIFF)
            )
        ]
        total_changes = len(diff_items)
        last_summary["total_changes"] = total_changes
        if total_changes > 0:
            total_numeric_diff = sum(d['numeric_diff'] for d in diff_items if d['numeric_diff'] is not None)
            filtered_percentage_diff = [d['percentage_diff'] for d in diff_items if d['percentage_diff'] is not None]
            avg_numeric = round(total_numeric_diff / total_changes, 2)
            last_summary["avg_numeric_diff"] = avg_numeric
            if filtered_percentage_diff:
                avg_percentage = round(sum(filtered_percentage_diff) / len(filtered_percentage_diff), 2)
                last_summary["avg_percentage_diff"] = avg_percentage

        flash("Changes saved successfully!", "success")
        return render_template('iteration_1.html', results=last_results, summary=last_summary)
    except Exception as e:
        print(f"Error while saving changes: {e}")
        flash("An error occurred while saving changes. Please try again.", "danger")
        return render_template('iteration_1.html', results=last_results, summary=last_summary)


@monitoring_mrch_bp.route('/export_excel_MRCH')
def export_excel_MRCH():
    if last_results is None:
        abort(400, "No data available for export")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comparison Results"

    # Формуємо унікальні заголовки
    base_headers = list(COLUMN_MAPPING.keys())
    extra_headers = ['price1', 'price2', 'numeric_diff', 'percentage_diff', 'photo_last_week', 'promo_last_week']
    headers = base_headers + [h for h in extra_headers if h not in base_headers]

    # Стилізація заголовків
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Додаємо заголовки
    sheet.append(headers)
    for col_num, column_title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Стиль для відхилень
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Додавання даних з перевіркою на порожні рядки
    for result in last_results:
        # Перевірка: якщо стовпець 'id' порожній, пропускаємо рядок
        if not result.get('id') or str(result.get('id')).strip() == '':
            continue

        # Генеруємо дані рядка
        row_data = [result.get(header, '') for header in headers]

        # Перевірка, чи рядок повністю порожній
        if all(cell == '' for cell in row_data):
            continue  # Пропускаємо рядок, якщо всі клітинки порожні

        # Додаємо дані до Excel
        sheet.append(row_data)

        # Виділення значних відхилень
        row_idx = sheet.max_row

        # Підсвітка для percentage_diff
        percentage_diff = result.get('percentage_diff')
        percentage_diff_col = headers.index('percentage_diff') + 1
        if percentage_diff is not None and (percentage_diff > 10 or percentage_diff < -10):
            sheet.cell(row=row_idx, column=percentage_diff_col).fill = red_fill

        # Підсвітка для numeric_diff
        numeric_diff = result.get('numeric_diff')
        numeric_diff_col = headers.index('numeric_diff') + 1
        if numeric_diff is not None and abs(numeric_diff) >= 10:
            sheet.cell(row=row_idx, column=numeric_diff_col).fill = red_fill

    # Збереження у файл
    excel_output = BytesIO()
    workbook.save(excel_output)
    excel_output.seek(0)

    response = make_response(excel_output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=exported_results.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return response


@monitoring_mrch_bp.route('/export_csv_MRCH')
def export_csv_MRCH():
    if last_results is None:
        abort(400, "No data available for export")

    si = StringIO()
    writer = csv.writer(si, quoting=csv.QUOTE_MINIMAL)

    base_headers = list(COLUMN_MAPPING.keys())
    extra_headers = ['price1', 'price2', 'numeric_diff', 'percentage_diff', 'promo', 'photo_last_week', 'promo_last_week']
    headers = base_headers + [h for h in extra_headers if h not in base_headers]

    writer.writerow(headers)

    for result in last_results:
        row = [result.get(header, '') for header in headers]
        writer.writerow(row)

    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=exported_results.csv"
    output.headers["Content-type"] = "text/csv"

    return output


if __name__ == '__main__':
    app.run(debug=True, port=8080)
